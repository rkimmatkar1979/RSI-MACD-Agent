"""
Streamlit dashboard for the Nifty 100 Swing Trading Agent.

Run with:  streamlit run app.py
"""

import html
import io
import re

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
import pandas as pd
import plotly.graph_objects as go
import streamlit as st
from plotly.subplots import make_subplots

import config
import db_handler
import fundamentals
from ai_analyst import get_ai_recommendations
from scheduler import is_market_open, run_pipeline
from strategy import generate_shortlist
from ta_engine import (
    calculate_buy_sell_pressure,
    calculate_volume_metrics,
    describe_macd_pattern,
    get_chart_data,
    get_reference_session,
    nearest_fib_level,
)

@st.cache_data(show_spinner=False)
def _df_to_styled_excel(df: pd.DataFrame) -> bytes:
    """Return a .xlsx file bytes with black cells and white text."""
    wb = openpyxl.Workbook()
    ws = wb.active
    black_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = black_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            if isinstance(value, (list, dict)):
                value = str(value)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = black_fill
            cell.font = Font(color="FFFFFF")
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


st.set_page_config(
    page_title=config.APP_TITLE,
    layout="wide",
    page_icon=config.APP_PAGE_ICON,
    initial_sidebar_state="expanded",
)

# Dark mode toggle - the widget itself lives in the sidebar (below). Its
# preference is mirrored into the plain "dark_mode_pref" key (via the
# toggle's on_change callback) rather than read directly from the widget's
# own "dark_mode" key, because the sidebar - and therefore the toggle - is
# never rendered during the "scan/analysis in progress" minimal pages below.
# A widget's session_state entry is dropped if the widget isn't instantiated
# on a run, which was silently resetting dark mode back to light once those
# pages finished. "dark_mode_pref" isn't tied to any widget, so it survives.
dark_mode = st.session_state.get("dark_mode_pref", False)

# Colour tokens shared by the CSS below, the Shortlist/Custom Analysis tables
# (pandas Styler), and the Plotly chart layouts - keeping a single
# light/dark palette here means every themed element stays in sync.
if dark_mode:
    COLOR_APP_BG = "#0E1117"
    COLOR_SECONDARY_BG = "#1C1F26"
    COLOR_TEXT = "#FAFAFA"
    COLOR_BORDER = "#FFFFFF"
    COLOR_TABLE_BG = "#FFFFFF"
    COLOR_TABLE_TEXT = "#000000"
    COLOR_TABLE_GRID = "#000000"
    PLOTLY_PAPER_BG = "#0E1117"
    PLOTLY_PLOT_BG = "#1C1F26"
    PLOTLY_FONT_COLOR = "#FAFAFA"
    PLOTLY_GRID_COLOR = "#3A3F4B"
    PLOTLY_ZERO_COLOR = "#6E7681"
    PLOTLY_ANNOTATION_BG = "rgb(30, 33, 40)"
else:
    COLOR_APP_BG = "#FAF6EC"
    COLOR_SECONDARY_BG = "#F2ECDD"
    COLOR_TEXT = "#000000"
    COLOR_BORDER = "#FFFFFF"
    COLOR_TABLE_BG = "#FFFFFF"
    COLOR_TABLE_TEXT = "#000000"
    COLOR_TABLE_GRID = "#000000"
    PLOTLY_PAPER_BG = "#FAF6EC"
    PLOTLY_PLOT_BG = "#FFFDF6"
    PLOTLY_FONT_COLOR = "#000000"
    PLOTLY_GRID_COLOR = "#D0D0D0"
    PLOTLY_ZERO_COLOR = "#A0A0A0"
    PLOTLY_ANNOTATION_BG = "rgba(255, 255, 255, 0.7)"

# Pull the title up to sit ~12px below Streamlit's top header bar (default
# header height ~3.75rem), and give every "boxed" UI element (tables, alerts,
# buttons, expanders, dropdowns, inputs) a flat border with square corners
# for a consistent look across the app.
_dark_mode_css = ""
if dark_mode:
    _dark_mode_css = f"""
    .stApp, [data-testid="stAppViewContainer"], [data-testid="stMain"],
    [data-testid="stHeader"], [data-testid="stBottomBlockContainer"] {{
        background-color: {COLOR_APP_BG} !important;
    }}
    [data-testid="stSidebar"] {{
        background-color: {COLOR_SECONDARY_BG} !important;
    }}
    .stApp p, .stApp span, .stApp label, .stApp div, .stApp li,
    .stApp h1, .stApp h2, .stApp h3, .stApp h4, .stApp h5, .stApp h6,
    [data-testid="stMetricValue"], [data-testid="stMetricLabel"],
    [data-testid="stMetricDelta"] {{
        color: {COLOR_TEXT} !important;
    }}
    [data-testid="stExpander"], [data-testid="stExpander"] summary,
    .stButton > button,
    [data-testid="stSegmentedControl"] label,
    div[data-baseweb="select"] > div,
    [data-testid="stTextInput"] input, [data-testid="stNumberInput"] input,
    [data-testid="stDataFrame"] {{
        background-color: {COLOR_SECONDARY_BG} !important;
    }}
    hr {{
        border-color: {COLOR_BORDER} !important;
        border-top-width: 2px !important;
        margin-top: 0.25rem !important;
        margin-bottom: 0.25rem !important;
    }}
    /* select/multiselect dropdown popovers render in a portal outside .stApp */
    [data-baseweb="popover"], [data-baseweb="menu"] {{
        background-color: {COLOR_SECONDARY_BG} !important;
    }}
    [data-baseweb="menu"] li, [data-baseweb="menu"] * {{
        color: {COLOR_TEXT} !important;
    }}
    /* Checkbox widget border/background in dark mode */
    [data-testid="stCheckbox"] span,
    [data-testid="stCheckbox"] input + div {{
        border-color: #FFFFFF !important;
        background-color: #000000 !important;
    }}
    [data-testid="stCheckbox"] label span {{
        color: {COLOR_TEXT} !important;
    }}
    /* Scrollbar visibility in dark mode */
    ::-webkit-scrollbar {{
        width: 8px;
        height: 8px;
    }}
    ::-webkit-scrollbar-track {{
        background: #1C1F26;
    }}
    ::-webkit-scrollbar-thumb {{
        background: #555555;
        border-radius: 4px;
    }}
    ::-webkit-scrollbar-thumb:hover {{
        background: #888888;
    }}
    """

st.markdown(
    f"""
    <style>
    *, html {{ scroll-behavior: smooth; }}
    .block-container {{ padding-top: calc(3.75rem + 12px); }}
    @keyframes shimmer {{
        0% {{ background-position: -1000px 0; }}
        100% {{ background-position: 1000px 0; }}
    }}
    .shimmer-line {{
        background: linear-gradient(90deg, {COLOR_SECONDARY_BG} 25%, {COLOR_TABLE_GRID} 50%, {COLOR_SECONDARY_BG} 75%);
        background-size: 1000px 100%;
        animation: shimmer 1.5s infinite linear;
        border-radius: 3px;
        margin-bottom: 10px;
    }}

    [data-testid="stDataFrame"],
    [data-testid="stAlert"],
    [data-testid="stExpander"],
    .stButton > button,
    [data-testid="stButtonGroup"] button,
    div[data-baseweb="select"] > div,
    [data-testid="stTextInput"] input,
    [data-testid="stNumberInput"] input {{
        border-radius: 0 !important;
        border: 1px solid {COLOR_BORDER} !important;
    }}

    /* Segmented control (tabs) - the "Segmented Control" testid doesn't
       actually exist in the rendered DOM (it's stButtonGroup under the
       hood), so target that directly for both the unselected and selected
       segment backgrounds/text. */
    [data-testid="stButtonGroup"] button {{
        background-color: {COLOR_SECONDARY_BG} !important;
    }}
    [data-testid="stButtonGroup"] button,
    [data-testid="stButtonGroup"] button * {{
        color: {COLOR_TEXT} !important;
    }}
    [data-testid="stBaseButton-primary"],
    [data-testid="stBaseButton-primary"] * {{
        color: white !important;
    }}
    [data-testid="stDownloadButton"] button {{
        background-color: #1565C0 !important;
        color: white !important;
        border: none !important;
    }}
    [data-testid="stDownloadButton"] button * {{
        color: white !important;
    }}
    [data-testid="stDownloadButton"] button:hover {{
        background-color: #3E1A00 !important;
        color: white !important;
    }}

    /* ---- Mobile (≤ 640px) -------------------------------------------- */
    @media (max-width: 640px) {{
        /* Tighten page padding on small screens */
        .block-container {{
            padding-left: 0.75rem !important;
            padding-right: 0.75rem !important;
        }}
        /* Stack every st.columns() layout vertically */
        [data-testid="stHorizontalBlock"] {{
            flex-wrap: wrap !important;
        }}
        [data-testid="stHorizontalBlock"] > [data-testid="stColumn"] {{
            width: 100% !important;
            flex: 1 1 100% !important;
            min-width: unset !important;
        }}
        /* Tab bar: scroll horizontally so all tabs stay on one line */
        [data-testid="stButtonGroup"] {{
            display: flex !important;
            flex-wrap: nowrap !important;
            overflow-x: auto !important;
            -webkit-overflow-scrolling: touch;
        }}
        /* Tables: horizontal scroll rather than overflow clip */
        [data-testid="stDataFrame"] > div {{
            overflow-x: auto !important;
        }}
        /* Heatmap: allow horizontal scroll on very narrow screens */
        .stPlotlyChart > div {{
            overflow-x: auto !important;
        }}
        /* Download button stretches to full width when column stacks */
        [data-testid="stDownloadButton"] {{
            width: 100% !important;
        }}
        [data-testid="stDownloadButton"] button {{
            width: 100% !important;
        }}
    }}
    {_dark_mode_css}
    </style>
    """,
    unsafe_allow_html=True,
)

# init_db() runs idempotent CREATE TABLE/ALTER checks - with a remote Turso
# database each check is a network round-trip, so this is cached process-wide
# (st.cache_resource) to run once per app deployment rather than once per
# browser session.
@st.cache_resource(show_spinner="Initializing database...")
def _ensure_db_initialized():
    db_handler.init_db()
    return True


_ensure_db_initialized()

st.title(config.APP_TITLE)
st.caption(config.APP_SUBTITLE)

# ---------------------------------------------------------------------------
# Authentication - Google sign-in via Streamlit's built-in auth (requires
# Authlib + a [auth]/[auth.google] section in .streamlit/secrets.toml, see
# .streamlit/secrets.toml.example). Toggle with config.AUTH_ENABLED.
# ---------------------------------------------------------------------------
if config.AUTH_ENABLED:
    if not getattr(st.user, "is_logged_in", False):
        st.write("Please sign in with your Google account to continue.")
        if st.button("🔐 Log in with Google", type="primary"):
            try:
                st.login("google")
            except Exception as e:
                st.error(
                    "Google sign-in isn't configured yet. Copy "
                    ".streamlit/secrets.toml.example to .streamlit/secrets.toml "
                    f"and fill in your Google OAuth credentials. ({e})"
                )
        st.stop()

    user_email = (st.user.email or "").lower()
    is_admin = user_email in config.AUTH_ADMIN_EMAILS

    # First-come-first-served access, capped at AUTH_MAX_USERS - admins are
    # exempt and can free up slots from the Admin tab below.
    if not is_admin:
        user_status = db_handler.get_user_status(user_email)

        if user_status == "revoked":
            st.error("Your access to this app has been revoked by the administrator.")
            if st.button("OK, sign me out"):
                st.logout()
            st.stop()

        if user_status != "active":
            if db_handler.get_authorized_user_count() >= config.AUTH_MAX_USERS:
                st.error(
                    f"This app is limited to {config.AUTH_MAX_USERS} users and that limit "
                    "has already been reached. Contact the administrator for access."
                )
                if st.button("Log out"):
                    st.logout()
                st.stop()
            db_handler.register_user(user_email, st.user.name or user_email)
else:
    user_email = ""
    is_admin = False

st.session_state["_user_email"] = user_email

# The Custom Analysis tab (below) is gated on this rather than is_admin alone
# - with auth disabled there's no concept of "other users" yet, so the tool
# is available to whoever is running the app; once AUTH_ENABLED is turned on
# it automatically locks down to AUTH_ADMIN_EMAILS only.
can_use_admin_tools = is_admin or not config.AUTH_ENABLED

# ---------------------------------------------------------------------------
# Run a fresh scan on demand - rendered as its own minimal page (just a
# progress bar, no sidebar/tabs/other widgets) so nothing else on screen can
# trigger a competing rerun that would cancel the scan partway through (e.g.
# switching tabs used to abort an in-progress scan).
# ---------------------------------------------------------------------------
if st.session_state.get("scan_in_progress"):
    st.title(config.APP_TITLE)
    st.caption(config.APP_SUBTITLE)
    st.markdown("---")
    st.subheader("🔄 Scanning Nifty 100...")
    st.caption(
        "Running full scan and requesting AI commentary — this takes a minute or two. "
        "The page will refresh automatically when done."
    )
    progress_bar = st.progress(0.0, text="Starting scan...")

    def _progress_cb(i, total, ticker):
        progress_bar.progress(i / total, text=f"Scanning {ticker} ({i}/{total})")

    try:
        shortlist, ai_commentary, scan_date = run_pipeline(progress_callback=_progress_cb)
        # The new scan's results would otherwise be hidden behind the 5-minute
        # cache on these DB reads (see db_handler) - clear it so this scan
        # shows up immediately below.
        db_handler.get_latest_scan.clear()
        db_handler.get_available_scan_dates.clear()
        db_handler.get_scan_by_date.clear()
        db_handler.get_scan_timestamps.clear()
        st.session_state.pop("date_select", None)
        db_handler.log_event(
            st.session_state.get("_user_email", ""),
            "scan_run",
            {"scan_date": scan_date, "shortlist_size": len(shortlist)},
        )
        st.session_state["scan_message"] = (
            "success",
            f"Scan complete for {scan_date}: {len(shortlist)} qualifying setup(s) found.",
        )
    except Exception as e:
        st.session_state["scan_message"] = (
            "error",
            f"Scan failed — please try again. If the issue persists, check your internet connection or LLM API key. (Detail: {e})",
        )
    finally:
        st.session_state["scan_in_progress"] = False
    st.rerun()

# ---------------------------------------------------------------------------
# Run a custom AI analysis on demand (Custom Analysis tab, below) - same
# minimal-page pattern as the full scan above, so it can't be cancelled by
# clicking elsewhere on the page while it runs.
# ---------------------------------------------------------------------------
if st.session_state.get("custom_analysis_in_progress"):
    custom_tickers = st.session_state.get("custom_analysis_request", [])
    st.title(config.APP_TITLE)
    st.caption(config.APP_SUBTITLE)
    st.markdown("---")
    st.subheader(f"🤖 Analysing {len(custom_tickers)} stock(s)...")
    st.caption(
        "Fetching technicals and generating AI commentary — this takes a minute or so. "
        "The page will refresh automatically when done."
    )
    progress_bar = st.progress(0.0, text="Starting analysis...")

    def _custom_progress_cb(i, total, ticker):
        progress_bar.progress(i / total, text=f"Analyzing {ticker} ({i}/{total})")

    try:
        custom_df = generate_shortlist(tickers=custom_tickers, progress_callback=_custom_progress_cb)
        progress_bar.progress(1.0, text="Fetching news and generating AI commentary...")
        with st.spinner("Fetching news and generating AI commentary - this can take a minute..."):
            custom_commentary = get_ai_recommendations(custom_df)
        db_handler.save_custom_analysis(custom_tickers, custom_df, custom_commentary)
        db_handler.get_available_custom_analyses.clear()
        db_handler.get_custom_analysis_by_id.clear()
        db_handler.log_event(
            st.session_state.get("_user_email", ""),
            "custom_analysis_run",
            {"tickers": custom_tickers, "ticker_count": len(custom_tickers)},
        )
    except Exception as e:
        st.session_state["custom_analysis_error"] = (
            f"Analysis failed — please try again. If the issue persists, check your internet connection or LLM API key. (Detail: {e})"
        )
    finally:
        st.session_state["custom_analysis_in_progress"] = False
        st.session_state.pop("custom_analysis_request", None)
    st.rerun()

if "scan_message" in st.session_state:
    level, message = st.session_state.pop("scan_message")
    getattr(st, level)(message)

# Load scan data before the sidebar so the date selector and scan-info
# label in the sidebar can reference it.
latest = db_handler.get_latest_scan()
available_dates = db_handler.get_available_scan_dates()
scan_timestamps = db_handler.get_scan_timestamps()

# ---------------------------------------------------------------------------
# Sidebar - controls
# ---------------------------------------------------------------------------
with st.sidebar:
    # Force readable text for everything in the sidebar - the default muted
    # caption color is low-contrast against the sidebar background in both
    # the light (beige) and dark themes.
    st.markdown(
        f"<style>[data-testid='stSidebar'] * {{ color: {COLOR_TEXT} !important; }}</style>",
        unsafe_allow_html=True,
    )

    st.header("Controls")

    st.toggle(
        "🌙 Dark mode",
        value=dark_mode,
        key="dark_mode",
        on_change=lambda: st.session_state.update(dark_mode_pref=st.session_state["dark_mode"]),
    )

    if dark_mode:
        st.info(config.DARK_MODE_NOTICE, icon=None)


    if config.AUTH_ENABLED:
        st.caption(f"Signed in as **{st.user.name or user_email}**" + (" 👑 (admin)" if is_admin else ""))
        if st.button("Log out", key="sidebar_logout"):
            st.logout()

        st.markdown("---")

    if st.button("🔍 Run Full Scan Now", type="primary", use_container_width=True):
        st.session_state["_pending_active_tab"] = st.session_state.get("active_tab")
        st.session_state["scan_in_progress"] = True
        st.rerun()

    st.markdown("---")

    try:
        if is_market_open():
            st.success("NSE market is currently OPEN")
        else:
            st.info("NSE market is currently CLOSED — scan will use the latest available daily candle.")
    except Exception:
        st.warning("Could not determine market hours.")

    if available_dates:
        st.markdown("---")
        selected_date = st.selectbox(
            "Viewing scan from:",
            available_dates,
            index=0,
            key="date_select",
            format_func=lambda d: scan_timestamps.get(d, d),
        )
    else:
        selected_date = None

    st.markdown("---")
    with st.expander("⚙️ Strategy Parameters", expanded=False):
        st.caption(
            f"Universe: {len(config.SCAN_UNIVERSE)} tickers "
            f"(Nifty 200 + Gold/Silver)"
        )
        st.caption(f"Fibonacci lookback: {config.FIB_LOOKBACK_DAYS} trading days (~3-4 months)")
        st.caption(f"RSI thresholds: oversold < {config.RSI_OVERSOLD}, overbought > {config.RSI_OVERBOUGHT}")
        st.caption(f"MACD: {config.MACD_FAST}/{config.MACD_SLOW}/{config.MACD_SIGNAL}")
        st.caption(
            f"Volume surge: >= {config.VOLUME_SURGE_RATIO}x its "
            f"{config.VOLUME_AVG_WINDOW}-day average"
        )
        st.caption(
            f"Sector trend: {config.SECTOR_TREND_LOOKBACK_DAYS}-day sector avg return, "
            f">= {config.SECTOR_TREND_THRESHOLD * 100:.1f}% for a score bonus"
        )
        st.caption(
            f"Buy/Sell pressure: {config.BUY_SELL_PRESSURE_WINDOW}-day volume split "
            "by up-days vs down-days (proxy, not live order-book data)"
        )

    st.markdown("---")
    st.markdown(
        f"Built by <a href='{config.AUTHOR_LINKEDIN_URL}' "
        f"target='_blank' style='color:#1a73e8;'>{config.AUTHOR_NAME}</a>",
        unsafe_allow_html=True,
    )

# ---------------------------------------------------------------------------
# Resolve selected scan (data loaded above, date picked in sidebar)
# ---------------------------------------------------------------------------
if latest is None:
    st.markdown("---")
    st.markdown("### Welcome to the Nifty 100 Swing Trading Agent")
    st.info(
        "No scans have been run yet. Click **Run Full Scan Now** in the sidebar to "
        "scan all Nifty 100 stocks for RSI/MACD + Fibonacci setups. The first scan "
        "takes 2–3 minutes."
    )
    st.markdown(
        "**What you'll get after the scan:**\n"
        "- 📋 **Shortlist** — stocks meeting the scoring threshold, ranked by signal strength\n"
        "- 🤖 **AI Commentary** — AI-generated write-up on each shortlisted stock\n"
        "- 📐 **Chart Analysis** — Fibonacci retracement levels + RSI/MACD chart for any stock\n"
        "- 🎯 **Custom Analysis** — on-demand AI write-up for any stocks you pick"
    )
    st.stop()

scan_date, ai_commentary, signals_df = latest

if selected_date and selected_date != scan_date:
    result = db_handler.get_scan_by_date(selected_date)
    if result is not None:
        scan_date, ai_commentary, signals_df = result

# ---------------------------------------------------------------------------
# Main views - split into tabs so the page isn't one long scroll. This also
# reads much better on mobile, where each tab's content fits the viewport on
# its own instead of stacking every section vertically.
# ---------------------------------------------------------------------------
selected_rows = []

TAB_SHORTLIST = "📋 Shortlist"
TAB_AI = "🤖 AI Commentary"
TAB_CHART = "📐 Chart Analysis"
TAB_CUSTOM = "🎯 Custom Analysis"
TAB_ADMIN = "👑 Admin"
_TAB_CONTAINER_KEYS = {
    TAB_SHORTLIST: "shortlist", TAB_AI: "ai", TAB_CHART: "chart",
    TAB_CUSTOM: "custom", TAB_ADMIN: "admin",
}

tab_names = [TAB_SHORTLIST, TAB_AI, TAB_CHART]
if can_use_admin_tools:
    tab_names.append(TAB_CUSTOM)
if is_admin:
    tab_names.append(TAB_ADMIN)

# st.tabs() doesn't persist the active tab across reruns triggered by a
# widget in a DIFFERENT tab (e.g. the Chart Analysis ticker selector) - it
# snaps back to the first tab. st.segmented_control is a normal
# session_state-backed widget, so it keeps whichever "tab" the user is on;
# each section below is rendered into its own container, and all but the
# active one are hidden via CSS.
# Restore the tab that was active before a "Run Full Scan Now" /
# "Get AI Analysis" cycle, if one just finished (see those button handlers).
if "_pending_active_tab" in st.session_state:
    _restored_tab = st.session_state.pop("_pending_active_tab")
    if _restored_tab in tab_names:
        st.session_state["active_tab"] = _restored_tab

if st.session_state.get("active_tab") not in tab_names:
    st.session_state["active_tab"] = tab_names[0]
_tab_col, _dm_topbar_col = st.columns([11, 1])
with _tab_col:
    active_tab = st.segmented_control(
        "View", tab_names, required=True, key="active_tab", label_visibility="collapsed",
    )
with _dm_topbar_col:
    st.toggle(
        "🌙",
        value=dark_mode,
        key="dark_mode_topbar",
        on_change=lambda: st.session_state.update(dark_mode_pref=st.session_state["dark_mode_topbar"]),
        help="Toggle dark mode",
        label_visibility="collapsed",
    )

_prev_tab = st.session_state.get("_last_logged_tab")
if active_tab != _prev_tab:
    st.session_state["_last_logged_tab"] = active_tab
    db_handler.log_event(user_email, "tab_switch", {"tab": active_tab})

# One combined <style> block (rather than one st.markdown call per hidden
# tab) - fewer elements for Streamlit to ship to the frontend on every rerun.
# Also fades the active tab's container in on each switch - a CSS animation
# replays whenever an element goes from display:none to visible, so this
# gives a smooth transition without changing the hide/show mechanism itself.
_hidden_selectors = ", ".join(
    f"div[class*='st-key-tab_{_TAB_CONTAINER_KEYS[name]}']"
    for name in tab_names if name != active_tab
)
_hide_rule = f"{_hidden_selectors} {{ display: none; }}" if _hidden_selectors else ""
st.markdown(
    f"""
    <style>
    {_hide_rule}
    @keyframes tabFadeIn {{ from {{ opacity: 0; }} to {{ opacity: 1; }} }}
    div[class*="st-key-tab_"] {{ animation: tabFadeIn 0.2s ease-in-out; }}
    @media (max-width: 768px) {{
        [data-testid="stHorizontalBlock"] {{ flex-wrap: wrap !important; }}
        [data-testid="stHorizontalBlock"] > [data-testid="column"] {{
            min-width: 45% !important;
            flex: 1 1 45% !important;
        }}
    }}
    </style>
    """,
    unsafe_allow_html=True,
)

tab_shortlist = st.container(key="tab_shortlist")
tab_ai = st.container(key="tab_ai")
tab_chart = st.container(key="tab_chart")
tab_custom = st.container(key="tab_custom") if can_use_admin_tools else None
tab_admin = st.container(key="tab_admin") if is_admin else None

# ---------------------------------------------------------------------------
# Shortlist table styling - color cues so strong/weak signals are visible at
# a glance instead of requiring a column-by-column read.
# ---------------------------------------------------------------------------
_MAX_SCORE = (
    config.SCORE_FIB_KEY_LEVEL + config.SCORE_RSI_EXTREME
    + config.SCORE_MACD_PROXIMITY + config.SCORE_VOLUME + config.SCORE_SECTOR_TREND
)


def _style_rsi(val):
    """Green = oversold (potential bullish reversal), red = overbought (bearish)."""
    if val <= config.RSI_OVERSOLD:
        return "background-color: #d4edda; color: #155724; font-weight: 600"
    if val >= config.RSI_OVERBOUGHT:
        return "background-color: #f8d7da; color: #721c24; font-weight: 600"
    return ""


def _style_score(val):
    """Shade Score green, more intensely the closer it is to the max of 100."""
    ratio = max(0.0, min(1.0, val / _MAX_SCORE))
    if ratio >= 0.7:
        return "background-color: #c3e6cb; color: #155724; font-weight: 700"
    if ratio >= 0.45:
        return "background-color: #e6f4ea; color: #1e7e34; font-weight: 600"
    return ""


def _style_macd_diff(val):
    """Color the MACD-Signal diff by sign: green = bullish, red = bearish."""
    try:
        diff = float(val.split(" ")[0])
    except (ValueError, IndexError):
        return ""
    if diff > 0:
        return "color: #2e8b57; font-weight: 600"
    if diff < 0:
        return "color: #c0392b; font-weight: 600"
    return ""


# ---------------------------------------------------------------------------
# Tab 1: Shortlist table
# ---------------------------------------------------------------------------
with tab_shortlist:
    st.subheader(f"Shortlist — {scan_date}")
    _scan_ts = scan_timestamps.get(scan_date, "")
    if _scan_ts:
        st.caption(f"Last updated: {_scan_ts}")

    with st.expander("📐 How the score is calculated (max 100)", expanded=False):
        st.markdown(
            f"""
| # | Component | Max points | Awarded when... |
|---|---|---|---|
| 1 | Fibonacci proximity | **{config.SCORE_FIB_KEY_LEVEL}** | Price is within {config.FIB_PROXIMITY_PCT * 100:.0f}% of a **key** level (50% / 61.8%) |
| 1b | (or) | **{config.SCORE_FIB_OTHER_LEVEL}** | ...or within {config.FIB_PROXIMITY_PCT * 100:.0f}% of any other level (0% / 23.6% / 38.2% / 100%) |
| 2 | RSI extreme | **{config.SCORE_RSI_EXTREME}** | RSI(14) <= {config.RSI_OVERSOLD} (oversold) or >= {config.RSI_OVERBOUGHT} (overbought) |
| 3 | MACD crossover proximity | **{config.SCORE_MACD_PROXIMITY}** | MACD histogram is small and shrinking - converging toward a crossover |
| 4 | Volume confirmation | **{config.SCORE_VOLUME}** | Latest volume >= {config.VOLUME_SURGE_RATIO}x its {config.VOLUME_AVG_WINDOW}-day average |
| 5 | Sector trend alignment | **{config.SCORE_SECTOR_TREND}** | Stock's bullish/bearish bias is confirmed by its sector's {config.SECTOR_TREND_LOOKBACK_DAYS}-day average return moving >= {config.SECTOR_TREND_THRESHOLD * 100:.1f}% the same direction |
| | **Total (best case)** | **{config.SCORE_FIB_KEY_LEVEL + config.SCORE_RSI_EXTREME + config.SCORE_MACD_PROXIMITY + config.SCORE_VOLUME + config.SCORE_SECTOR_TREND}** | 1 (key level) + 2 + 3 + 4 + 5 |

Items 1/1b are mutually exclusive (only the nearest Fib level counts). Item 5
is intentionally a small/secondary factor - sector trend never decides a
setup on its own, it only adds a bit of conviction when it agrees with the
stock's own signals.
            """
        )

    if signals_df.empty:
        st.info("No stocks met the scoring threshold on this date.")
    else:
        # --- Sector trend heatmap -------------------------------------------
        # One cell per sector represented in this shortlist (not the full
        # scan universe) - a quick read on which sectors are rotating in/out
        # of favour, using the same sector_trend_pct fed into Score component 5.
        st.markdown("**Sector trend snapshot**")
        sector_trend_df = (
            signals_df[["sector", "sector_trend_pct"]]
            .drop_duplicates(subset="sector")
            .sort_values("sector_trend_pct", ascending=False)
        )
        sector_names = sector_trend_df["sector"].tolist()
        sector_values = (sector_trend_df["sector_trend_pct"] * 100).round(2).tolist()
        heatmap_fig = go.Figure(data=go.Heatmap(
            z=[sector_values],
            x=sector_names,
            y=[""],
            colorscale="RdYlGn",
            zmid=0,
            text=[[f"{v:+.2f}%" for v in sector_values]],
            texttemplate="%{text}",
            hovertemplate="%{x}: %{z:+.2f}%<extra></extra>",
            colorbar=dict(title="%", thickness=12, len=0.7),
        ))
        heatmap_fig.update_layout(
            height=150,
            margin=dict(l=10, r=10, t=10, b=30),
            yaxis=dict(showticklabels=False),
            paper_bgcolor=PLOTLY_PAPER_BG,
            plot_bgcolor=PLOTLY_PLOT_BG,
            font_color=PLOTLY_FONT_COLOR,
        )
        st.plotly_chart(heatmap_fig, use_container_width=True)
        st.caption(
            f"Average {config.SECTOR_TREND_LOOKBACK_DAYS}-day return of all scanned "
            "stocks in each sector represented in this shortlist - green = sector "
            "trending up, red = trending down (see Score component 5 above)."
        )

        # --- Day-over-day diff vs the previous retained scan ----------------
        # SCAN_HISTORY_RETENTION_DAYS bounds how far back `available_dates`
        # goes, so "previous" here means the next-older retained scan, not
        # necessarily yesterday.
        prev_scan_date = None
        prev_signals_df = None
        if scan_date in available_dates:
            _scan_idx = available_dates.index(scan_date)
            if _scan_idx + 1 < len(available_dates):
                prev_scan_date = available_dates[_scan_idx + 1]
                _prev_result = db_handler.get_scan_by_date(prev_scan_date)
                if _prev_result is not None:
                    _, _, prev_signals_df = _prev_result

        if prev_signals_df is not None and not prev_signals_df.empty:
            prev_scores = dict(zip(prev_signals_df["ticker"], prev_signals_df["score"]))
        else:
            prev_scores = None

        def _score_delta_display(row):
            if prev_scores is None:
                return "—"
            if row["ticker"] not in prev_scores:
                return "🆕 New"
            delta = int(row["score"]) - int(prev_scores[row["ticker"]])
            if delta > 0:
                return f"▲ +{delta}"
            if delta < 0:
                return f"▼ {delta}"
            return "→ 0"

        score_history = db_handler.get_score_history(
            tuple(signals_df["ticker"].tolist()), scan_date
        )

        display_df = signals_df.copy()
        display_df["score_delta_display"] = display_df.apply(_score_delta_display, axis=1)
        def _fmt_score_trend(scores):
            if not scores:
                return "—"
            recent = scores[-4:]
            parts = " → ".join(str(int(s)) for s in recent)
            if len(recent) >= 2:
                if recent[-1] > recent[0]:
                    return parts + " ▲"
                if recent[-1] < recent[0]:
                    return parts + " ▼"
                return parts + " →"
            return parts

        display_df["score_trend"] = display_df["ticker"].map(
            lambda t: _fmt_score_trend(score_history.get(t, []))
        )
        display_df["macd_hist_display"] = display_df.apply(
            lambda r: f"{r['macd_hist']:.2f} ({r['macd_hist_direction']})", axis=1
        )
        display_df["buy_sell_display"] = display_df.apply(
            lambda r: f"{r['buy_pct']:.0f}% / {r['sell_pct']:.0f}%", axis=1
        )
        display_df["sector_trend_display"] = (display_df["sector_trend_pct"] * 100).round(2)
        display_df["signals_display"] = display_df["reasons"].apply(
            lambda r: "; ".join(r) if r else "-"
        )

        display_df = display_df[[
            "ticker", "sector", "close", "rsi", "macd_hist_display", "nearest_fib_level",
            "nearest_fib_price", "fib_distance_pct", "week52_high",
            "pct_from_52w_high", "volume_ratio", "buy_sell_display",
            "sector_trend_display", "score", "score_delta_display", "score_trend",
            "signals_display",
        ]].rename(columns={
            "ticker": "Ticker",
            "sector": "Sector",
            "close": "Price",
            "rsi": "RSI",
            "macd_hist_display": "MACD-Signal Diff (dir)",
            "nearest_fib_level": "Nearest Fib",
            "nearest_fib_price": "Fib Price",
            "fib_distance_pct": "Fib Dist %",
            "week52_high": "52W High",
            "pct_from_52w_high": "% From 52W High",
            "volume_ratio": "Vol vs 20D Avg",
            "buy_sell_display": "Buy % / Sell %",
            "sector_trend_display": "Sector Trend % (10D)",
            "score": "Score",
            "score_delta_display": "Δ vs Prev",
            "score_trend": "Score Trend",
            "signals_display": "Signals",
        })
        display_df["Fib Dist %"] = (display_df["Fib Dist %"] * 100).round(2)
        display_df["% From 52W High"] = (display_df["% From 52W High"] * 100).round(2)
        display_df["Vol vs 20D Avg"] = display_df["Vol vs 20D Avg"].round(2)
        display_df[["Price", "RSI", "Fib Price", "52W High"]] = (
            display_df[["Price", "RSI", "Fib Price", "52W High"]].round(2)
        )

        # Compact view by default (better on mobile / narrow screens) - the
        # remaining columns are still available via the checkbox below, and
        # always shown in the per-row detail panel when a row is clicked.
        compact_columns = [
            "Ticker", "Sector", "Price", "RSI", "MACD-Signal Diff (dir)",
            "Nearest Fib", "Fib Dist %", "Score", "Δ vs Prev", "Score Trend",
        ]
        _search = st.text_input(
            "🔍 Filter by ticker", key="shortlist_search", placeholder="e.g. RELIANCE",
            label_visibility="collapsed",
        ).strip().upper()
        if _search:
            display_df = display_df[display_df["Ticker"].str.contains(_search, na=False)]

        col_check, col_download = st.columns([3, 1])
        with col_check:
            show_all_cols = st.checkbox(
                "Show all columns (52W high, volume, buy/sell pressure, sector trend)",
                value=False,
            )
        with col_download:
            st.download_button(
                "📥 Download Excel",
                data=_df_to_styled_excel(display_df),
                file_name=f"nifty100_shortlist_{scan_date}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                width="stretch",
            )
        table_df = display_df if show_all_cols else display_df[compact_columns]

        # With all columns shown, let the table keep its natural (wider)
        # width and scroll horizontally instead of squeezing every column
        # into the container.
        table_width = "content" if show_all_cols else "stretch"
        styled_table = (
            table_df.style.set_properties(**{
                "background-color": COLOR_TABLE_BG,
                "color": COLOR_TABLE_TEXT,
                "border": f"1px solid {COLOR_TABLE_GRID}",
            })
            .map(_style_rsi, subset=["RSI"])
            .map(_style_score, subset=["Score"])
            .map(_style_macd_diff, subset=["MACD-Signal Diff (dir)"])
        )
        select_event = st.dataframe(
            styled_table, width=table_width, hide_index=True,
            on_select="rerun", selection_mode="single-row", key="shortlist_table",
        )
        with st.expander("ℹ️ Reading the table", expanded=False):
            st.markdown(
                f"**RSI** — 🟩 green when oversold (≤ {config.RSI_OVERSOLD}, possible bullish reversal), "
                f"🟥 red when overbought (≥ {config.RSI_OVERBOUGHT}, possible bearish reversal).\n\n"
                f"**Score** — shaded green for stronger setups (darker = closer to max {_MAX_SCORE}). "
                f"**Δ vs Prev** — score change vs the previous retained scan"
                f"{f' ({prev_scan_date})' if prev_scan_date else ''}; 🆕 = new to the shortlist. "
                f"**Score Trend** — scores across the last {config.SCAN_HISTORY_RETENTION_DAYS} retained scans (oldest → newest), with ▲/▼/→ showing the overall direction.\n\n"
                "**MACD-Signal Diff (dir)** — positive = MACD above Signal line (bullish), negative = below (bearish). "
                "**(up)**/**(down)** shows whether the histogram rose or fell vs the prior session.\n\n"
                "**Fib Dist %** — how close price is to its nearest Fibonacci retracement level. "
                "Click a row for the full breakdown including 52W high, volume, buy/sell pressure, and sector trend."
            )
            if show_all_cols:
                st.markdown(
                    f"**% From 52W High** — how far current price sits below the 52-week high (0% = at the high). "
                    f"**Vol vs 20D Avg** — latest volume as a multiple of the {config.VOLUME_AVG_WINDOW}-day average "
                    f"(≥ {config.VOLUME_SURGE_RATIO}x = surge). "
                    f"**Buy % / Sell %** — volume-weighted proxy over the last {config.BUY_SELL_PRESSURE_WINDOW} sessions "
                    "(not live order-book data). "
                    "**Sector Trend % (10D)** — sector's average return over the last 10 sessions."
                )

        if prev_scores is not None:
            dropped_tickers = set(prev_scores) - set(signals_df["ticker"])
            if dropped_tickers:
                dropped_list = ", ".join(
                    f"{t} ({prev_scores[t]}/100)" for t in sorted(dropped_tickers)
                )
                st.caption(f"📉 Dropped from the shortlist since {prev_scan_date}: {dropped_list}")

        selected_rows = select_event["selection"]["rows"] if select_event else []
        if selected_rows:
            sel_row = signals_df.iloc[selected_rows[0]]
            st.markdown(f"**{sel_row['ticker']}** ({sel_row['sector']}) — score {sel_row['score']}/100")

            d1, d2, d3, d4, d5 = st.columns(5)
            with d1:
                st.metric(
                    "Nearest Fib",
                    f"{sel_row['nearest_fib_level']} @ {sel_row['nearest_fib_price']:.2f}",
                )
            with d2:
                st.metric(
                    "52W High", f"₹{sel_row['week52_high']:.2f}",
                    delta=f"-{sel_row['pct_from_52w_high'] * 100:.1f}%",
                )
            with d3:
                st.metric("Vol vs 20D Avg", f"{sel_row['volume_ratio']:.2f}x")
            with d4:
                st.metric("Buy % / Sell %", f"{sel_row['buy_pct']:.0f}% / {sel_row['sell_pct']:.0f}%")
            with d5:
                st.metric(
                    f"Sector Trend ({config.SECTOR_TREND_LOOKBACK_DAYS}D)",
                    f"{sel_row['sector_trend_pct'] * 100:+.2f}%",
                )

            for reason in sel_row["reasons"]:
                st.markdown(f"- {reason}")
        else:
            st.caption("👆 Click a row above to see its full signal breakdown.")

# ---------------------------------------------------------------------------
# Tab 2 helpers
# ---------------------------------------------------------------------------
def _split_ai_commentary(text, tickers):
    """Split the full AI commentary blob into {ticker: section_text} by ### TICKER headings."""
    if not text:
        return {}
    sections = {}
    matches = list(re.finditer(r"^###\s+(.+?)$", text, re.MULTILINE))
    for i, match in enumerate(matches):
        heading = match.group(1).strip()
        start = match.end()
        end = matches[i + 1].start() if i + 1 < len(matches) else len(text)
        section_text = text[start:end].strip()
        for ticker in tickers:
            base = ticker.replace(".NS", "").replace(".BO", "")
            if base in heading or ticker in heading:
                sections[ticker] = section_text
                break
    return sections


def _render_company_basics(basics):
    """Render key ratios, quarterly P&L, cash flow, and shareholding from a fundamentals dict."""
    info = basics["info"]
    q_pl = basics["quarterly_pl"]
    q_cf = basics["quarterly_cashflow"]
    major_holders = basics["major_holders"]

    def _v(key, fmt="{:.2f}"):
        val = info.get(key)
        if val is None:
            return "—"
        try:
            return fmt.format(float(val))
        except Exception:
            return "—"

    def _pct(key):
        val = info.get(key)
        if val is None:
            return "—"
        try:
            return f"{float(val) * 100:.1f}%"
        except Exception:
            return "—"

    def _mcap(key):
        val = info.get(key)
        if val is None:
            return "—"
        try:
            cr = float(val) / 1e7
            if cr >= 1e5:
                return f"₹{cr / 1e5:.1f}L Cr"
            if cr >= 1e3:
                return f"₹{cr / 1e3:.0f}K Cr"
            return f"₹{cr:.0f} Cr"
        except Exception:
            return "—"

    def _fmt_qtr_col(col):
        try:
            return col.strftime("%b '%y")
        except Exception:
            return str(col)

    def _fmt_cr_df(df, row_labels):
        available = [r for r in row_labels if r in df.index]
        if not available:
            return None
        out = df.loc[available].iloc[:, :4].copy()
        out.columns = [_fmt_qtr_col(c) for c in out.columns]
        out = out.apply(pd.to_numeric, errors="coerce") / 1e7
        out = out.round(0)
        return out

    # --- Key ratios (2 rows of 4 metrics) ---
    r1c1, r1c2, r1c3, r1c4 = st.columns(4)
    r1c1.metric("Market Cap", _mcap("marketCap"))
    r1c2.metric("P/E (TTM)", _v("trailingPE"))
    r1c3.metric("P/B", _v("priceToBook"))
    r1c4.metric("EPS (TTM)", _v("trailingEps"))

    r2c1, r2c2, r2c3, r2c4 = st.columns(4)
    r2c1.metric("ROE", _pct("returnOnEquity"))
    r2c2.metric("D/E Ratio", _v("debtToEquity"))
    r2c3.metric("Profit Margin", _pct("profitMargins"))
    r2c4.metric("Dividend Yield", _pct("dividendYield"))

    # --- Quarterly P&L ---
    pl_display = _fmt_cr_df(q_pl, fundamentals._PL_ROWS)
    if pl_display is not None:
        st.markdown("**Quarterly P&L** *(₹ Cr)*")
        st.dataframe(pl_display, use_container_width=True)
    else:
        st.caption("Quarterly P&L not available.")

    # --- Quarterly Cash Flow ---
    cf_display = _fmt_cr_df(q_cf, fundamentals._CF_ROWS)
    if cf_display is not None:
        st.markdown("**Quarterly Cash Flow** *(₹ Cr)*")
        st.dataframe(cf_display, use_container_width=True)
    else:
        st.caption("Cash flow data not available.")

    # --- Shareholding Pattern ---
    if major_holders is not None and not major_holders.empty:
        st.markdown("**Shareholding Pattern**")
        st.dataframe(major_holders, use_container_width=True)
    else:
        st.caption("Shareholding data not available.")


# ---------------------------------------------------------------------------
# Tab 2: AI commentary
# ---------------------------------------------------------------------------
with tab_ai:
    st.subheader("🤖 AI Analyst Commentary")
    _displayed_commentary = st.session_state.get(f"regen_{scan_date}", ai_commentary)
    if _displayed_commentary:
        _btn_col, _regen_col, _ = st.columns([1, 1, 6])
        with _btn_col:
            _escaped_comm = html.escape(_displayed_commentary)
            st.markdown(
                f'<textarea id="_ai_comm_text" style="position:fixed;left:-9999px">{_escaped_comm}</textarea>'
                '<button onclick="navigator.clipboard.writeText(document.getElementById(\'_ai_comm_text\').value)'
                '.then(()=>{{this.innerHTML=\'✅ Copied!\';setTimeout(()=>this.innerHTML=\'📋 Copy\',1500)}})"'
                ' style="padding:4px 12px;border-radius:4px;border:1px solid #ccc;cursor:pointer;background:transparent">📋 Copy</button>',
                unsafe_allow_html=True,
            )
        with _regen_col:
            if st.button("🔄 Regenerate", key="regen_ai_btn"):
                with st.spinner("Regenerating AI commentary..."):
                    try:
                        _new_comm = get_ai_recommendations(signals_df)
                        st.session_state[f"regen_{scan_date}"] = _new_comm
                        st.rerun()
                    except Exception as _e:
                        st.error(f"Regeneration failed: {_e}")
        # Prefetch all fundamentals in parallel when this tab is visible (cached, runs once per scan)
        if active_tab == TAB_AI:
            _prefetch_key = f"_basics_prefetched_{scan_date}"
            if not st.session_state.get(_prefetch_key):
                with st.spinner("Loading company fundamentals..."):
                    fundamentals.prefetch_all(signals_df["ticker"].tolist())
                st.session_state[_prefetch_key] = True

        _ai_sections = _split_ai_commentary(_displayed_commentary, signals_df["ticker"].tolist())

        for _, _ai_row in signals_df.iterrows():
            _ticker = _ai_row["ticker"]
            st.markdown("---")
            st.markdown(f"### {_ticker} &nbsp; <span style='font-size:0.8em;font-weight:400'>{_ai_row['sector']} | Score {_ai_row['score']}/100</span>", unsafe_allow_html=True)
            _section = _ai_sections.get(_ticker, "")
            if _section:
                st.markdown(_section)
            else:
                st.caption("AI commentary not parsed for this stock.")
            with st.expander("📊 Company Basics", expanded=False):
                _basics = fundamentals.get_company_basics(_ticker)
                if _basics:
                    _render_company_basics(_basics)
                else:
                    st.caption("Could not load fundamental data for this stock.")
    else:
        st.markdown("_No commentary available. Run a scan to generate AI commentary._")

# ---------------------------------------------------------------------------
# Tab 3: Fibonacci retracement analysis
# ---------------------------------------------------------------------------
with tab_chart:
    st.subheader("📐 Fibonacci Retracement Analysis")

    if not signals_df.empty:
        ticker_list = signals_df["ticker"].tolist()

        # Sync with a clicked shortlist row above, if any - clicking a different
        # row re-points this selector (and therefore the chart below) at that
        # ticker. Manual changes to the selector below still work independently.
        if selected_rows:
            sel_ticker = signals_df.iloc[selected_rows[0]]["ticker"]
            if sel_ticker in ticker_list:
                st.session_state["chart_ticker_select"] = sel_ticker

        if st.session_state.get("chart_ticker_select") not in ticker_list:
            st.session_state["chart_ticker_select"] = ticker_list[0]

        st.caption(
            "👆 Click a row in the **Shortlist** tab to load that stock's "
            "chart here, or pick one manually."
        )
        chart_ticker = st.selectbox(
            "Select a stock to analyze", ticker_list, key="chart_ticker_select"
        )

        _prev_chart = st.session_state.get("_last_logged_chart")
        if chart_ticker != _prev_chart:
            st.session_state["_last_logged_chart"] = chart_ticker
            db_handler.log_event(user_email, "chart_viewed", {"ticker": chart_ticker, "scan_date": scan_date})

        chart_df, levels, peak, trough = get_chart_data(chart_ticker)
        if chart_df is None:
            st.error(f"Could not load price data for {chart_ticker}.")
        else:
            # The actual swing high/low bars the Fib levels were derived from.
            fib_window = chart_df.tail(config.FIB_LOOKBACK_DAYS)
            peak_date = fib_window["High"].idxmax()
            trough_date = fib_window["Low"].idxmin()

            current_price = float(chart_df["Close"].iloc[-1])
            current_ratio = (peak - current_price) / (peak - trough) if peak != trough else float("nan")
            level_name, level_price, distance_pct = nearest_fib_level(current_price, levels)
            _, avg_volume_20, volume_ratio = calculate_volume_metrics(chart_df)

            # Limit the plotted chart to the most recent CHART_DISPLAY_MONTHS -
            # indicators above (RSI/MACD/Fib levels/swing high-low) are still
            # computed from the full history / FIB_LOOKBACK_DAYS window and
            # apply across this shorter view.
            display_cutoff = chart_df.index.max() - pd.DateOffset(months=config.CHART_DISPLAY_MONTHS)
            chart_display_df = chart_df[chart_df.index >= display_cutoff]

            # --- Price + volume + MACD chart ---------------------------------------
            fig = make_subplots(
                rows=3, cols=1, shared_xaxes=True,
                row_heights=[0.5, 0.15, 0.35], vertical_spacing=0.03,
                specs=[[{}], [{"secondary_y": True}], [{}]],
            )

            fig.add_trace(go.Candlestick(
                x=chart_display_df.index,
                open=chart_display_df["Open"], high=chart_display_df["High"],
                low=chart_display_df["Low"], close=chart_display_df["Close"],
                name=chart_ticker,
            ), row=1, col=1)

            # Shade the bands between consecutive Fibonacci levels so the
            # retracement grid reads as zones, not just lines. Opacity bumped
            # up to 0.18 (from an original 0.10) for better visibility against
            # the light beige chart background.
            zone_colors = [
                "rgba(128,128,128,0.18)", "rgba(224,123,57,0.18)",
                "rgba(184,134,11,0.18)", "rgba(46,139,87,0.18)",
                "rgba(31,119,180,0.18)", "rgba(128,128,128,0.18)",
            ]
            sorted_levels = sorted(levels.items(), key=lambda kv: kv[1])
            for i in range(len(sorted_levels) - 1):
                (_, y0), (_, y1) = sorted_levels[i], sorted_levels[i + 1]
                fig.add_hrect(
                    y0=y0, y1=y1, fillcolor=zone_colors[i % len(zone_colors)],
                    line_width=0, row=1, col=1,
                )

            # Darkgoldenrod (#b8860b) replaces the original gold (#d4af37) for
            # the 38.2% level - the original was low-contrast on a light
            # background.
            level_colors = {
                "0.0%": "grey", "23.6%": "#e07b39", "38.2%": "#b8860b",
                "50.0%": "#2e8b57", "61.8%": "#1f77b4", "100.0%": "grey",
            }
            for name, price in levels.items():
                color = level_colors.get(name, "grey")
                fig.add_hline(
                    y=price,
                    line_dash="dash",
                    line_color=color,
                    line_width=1.5,
                    annotation_text=f"{name}: {price:.2f}",
                    annotation_position="right",
                    annotation_font=dict(size=11, color=color),
                    annotation_bgcolor=PLOTLY_ANNOTATION_BG,
                    annotation_bordercolor=color,
                    annotation_borderwidth=1,
                    row=1, col=1,
                )

            fig.add_hline(
                y=current_price,
                line_dash="solid",
                line_color="#FF00FF",
                line_width=2.5,
                annotation_text=f"CMP: {current_price:.2f}",
                annotation_position="left",
                annotation_font=dict(size=12, color="#FF00FF"),
                annotation_bgcolor=PLOTLY_ANNOTATION_BG,
                annotation_bordercolor="#FF00FF",
                annotation_borderwidth=1,
                row=1, col=1,
            )

            # Mark the exact swing-high/swing-low bars the levels were measured
            # from - but only if they fall within the displayed window, since
            # the FIB_LOOKBACK_DAYS window (used to derive the levels) can
            # extend further back than the CHART_DISPLAY_MONTHS shown here.
            swing_x, swing_y, swing_text, swing_colors, swing_symbols = [], [], [], [], []
            if peak_date >= display_cutoff:
                swing_x.append(peak_date)
                swing_y.append(peak)
                swing_text.append(f"Swing High {peak:.2f}")
                swing_colors.append("#2e8b57")
                swing_symbols.append("triangle-down")
            if trough_date >= display_cutoff:
                swing_x.append(trough_date)
                swing_y.append(trough)
                swing_text.append(f"Swing Low {trough:.2f}")
                swing_colors.append("#c0392b")
                swing_symbols.append("triangle-up")

            if swing_x:
                fig.add_trace(go.Scatter(
                    x=swing_x,
                    y=swing_y,
                    mode="markers+text",
                    marker=dict(size=12, color=swing_colors, symbol=swing_symbols),
                    text=swing_text,
                    textposition="top center",
                    name="Swing points",
                    showlegend=False,
                ), row=1, col=1)

            # Volume bars (green/red by up/down day) + RSI line on a secondary axis.
            vol_colors = [
                "#2e8b57" if c >= o else "#c0392b"
                for c, o in zip(chart_display_df["Close"], chart_display_df["Open"])
            ]
            fig.add_trace(go.Bar(
                x=chart_display_df.index, y=chart_display_df["Volume"],
                marker_color=vol_colors, name="Volume", showlegend=False,
            ), row=2, col=1)
            fig.add_trace(go.Scatter(
                x=chart_display_df.index, y=chart_display_df["RSI"], mode="lines",
                line=dict(color="#6a3d9a", width=1.5), name="RSI (14)",
            ), row=2, col=1, secondary_y=True)
            fig.add_hline(
                y=config.RSI_OVERBOUGHT, line_dash="dot", line_color="#c0392b",
                line_width=1, row=2, col=1, secondary_y=True,
            )
            fig.add_hline(
                y=config.RSI_OVERSOLD, line_dash="dot", line_color="#2e8b57",
                line_width=1, row=2, col=1, secondary_y=True,
            )

            # MACD panel: histogram (green/red by sign) + MACD/Signal lines + zero line.
            macd_hist_colors = [
                "#2e8b57" if v >= 0 else "#c0392b" for v in chart_display_df["MACD_HIST"]
            ]
            fig.add_trace(go.Bar(
                x=chart_display_df.index, y=chart_display_df["MACD_HIST"],
                marker_color=macd_hist_colors, name="MACD Histogram", showlegend=False,
            ), row=3, col=1)
            fig.add_trace(go.Scatter(
                x=chart_display_df.index, y=chart_display_df["MACD"], mode="lines",
                line=dict(color="#1f77b4", width=1.5), name="MACD",
            ), row=3, col=1)
            fig.add_trace(go.Scatter(
                x=chart_display_df.index, y=chart_display_df["MACD_SIGNAL"], mode="lines",
                line=dict(color="#e07b39", width=1.5), name="Signal",
            ), row=3, col=1)
            fig.add_hline(y=0, line_color="grey", line_width=1, row=3, col=1)

            fig.update_layout(
                title=(
                    f"{chart_ticker} — Last {config.CHART_DISPLAY_MONTHS} Months "
                    f"(Fibonacci levels from {config.FIB_LOOKBACK_DAYS}-Day range) "
                    "+ Volume/RSI + MACD"
                ),
                xaxis_rangeslider_visible=False,
                height=900,
                yaxis_title="Price (INR)",
                yaxis2_title="Volume",
                yaxis3_title="MACD",
                legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
                paper_bgcolor=PLOTLY_PAPER_BG,
                plot_bgcolor=PLOTLY_PLOT_BG,
                font_color=PLOTLY_FONT_COLOR,
            )
            fig.update_xaxes(gridcolor=PLOTLY_GRID_COLOR, zerolinecolor=PLOTLY_ZERO_COLOR, tickfont=dict(color=PLOTLY_FONT_COLOR))
            fig.update_yaxes(gridcolor=PLOTLY_GRID_COLOR, zerolinecolor=PLOTLY_ZERO_COLOR, tickfont=dict(color=PLOTLY_FONT_COLOR))
            fig.update_yaxes(title_text="RSI", range=[0, 100], row=2, col=1, secondary_y=True)
            fig.update_xaxes(title_text="Date", tickformat="%d %b %Y", row=3, col=1)
            st.plotly_chart(fig, use_container_width=True)

            # --- Metrics row 1, with arrows showing current MACD movement -----------
            macd_now = float(chart_df["MACD"].iloc[-1])
            macd_prev = float(chart_df["MACD"].iloc[-2])
            hist_now = float(chart_df["MACD_HIST"].iloc[-1])
            hist_prev = float(chart_df["MACD_HIST"].iloc[-2])

            col1, col2, col3, col4, col5 = st.columns(5)
            with col1:
                st.metric("RSI (14)", f"{chart_df['RSI'].iloc[-1]:.1f}")
            with col2:
                st.metric("MACD Line", f"{macd_now:.3f}", delta=f"{macd_now - macd_prev:+.3f}")
            with col3:
                st.metric("MACD Histogram", f"{hist_now:.3f}", delta=f"{hist_now - hist_prev:+.3f}")
            with col4:
                st.metric(f"Volume vs {config.VOLUME_AVG_WINDOW}D Avg", f"{volume_ratio:.2f}x")
            with col5:
                st.metric("Last Close", f"₹{current_price:.2f}")

            st.caption(
                "Arrows on **MACD Line** / **MACD Histogram** show the change vs. the "
                "previous session (green = rising, red = falling) - a quick read on "
                "the current direction of MACD momentum."
            )

            # --- Metrics row 2: previous session, buy/sell pressure, sector ---------
            prev_date, prev_open, prev_close = get_reference_session(chart_df)
            buy_pct, sell_pct = calculate_buy_sell_pressure(chart_df)

            chart_match = signals_df[signals_df["ticker"] == chart_ticker]
            if not chart_match.empty:
                sector = chart_match.iloc[0]["sector"]
                sector_trend_pct = chart_match.iloc[0]["sector_trend_pct"]
            else:
                sector = config.SECTOR_MAP.get(chart_ticker, "Unknown")
                sector_trend_pct = float("nan")

            col6, col7, col8, col9, col10 = st.columns(5)
            with col6:
                st.metric("Prev Session Open", f"₹{prev_open:.2f}", help=f"Session: {prev_date}")
            with col7:
                st.metric(
                    "Prev Session Close", f"₹{prev_close:.2f}",
                    delta=f"{prev_close - prev_open:+.2f}", help=f"Session: {prev_date}",
                )
            with col8:
                st.metric("Buy % / Sell %", f"{buy_pct:.0f}% / {sell_pct:.0f}%")
            with col9:
                st.metric("Sector", sector)
            with col10:
                if not pd.isna(sector_trend_pct):
                    st.metric(f"Sector Trend ({config.SECTOR_TREND_LOOKBACK_DAYS}D)", f"{sector_trend_pct * 100:+.2f}%")
                else:
                    st.metric(f"Sector Trend ({config.SECTOR_TREND_LOOKBACK_DAYS}D)", "n/a")

            if is_market_open():
                st.caption(
                    f"**Prev Session** ({prev_date}): last completed session — today's live bar is excluded. "
                    f"**Buy % / Sell %**: volume-weighted proxy over {config.BUY_SELL_PRESSURE_WINDOW} sessions, not live order-book data."
                )
            else:
                st.caption(
                    f"**Prev Session** ({prev_date}): today's just-closed session. "
                    f"**Buy % / Sell %**: volume-weighted proxy over {config.BUY_SELL_PRESSURE_WINDOW} sessions, not live order-book data."
                )

            st.markdown("**MACD Pattern Analysis**")
            st.info(describe_macd_pattern(chart_df))

            # --- Fibonacci retracement breakdown -------------------------------------
            st.markdown("**How These Fibonacci Levels Were Chosen**")
            st.caption(
                f"Swing High: ₹{peak:.2f} on {peak_date.date()} | "
                f"Swing Low: ₹{trough:.2f} on {trough_date.date()} "
                f"(highest High / lowest Low over the last {config.FIB_LOOKBACK_DAYS} "
                "trading days, marked on the chart above). Each level = Swing High − "
                "ratio × (Swing High − Swing Low)."
            )

            fib_rows = []
            for name, price in sorted(levels.items(), key=lambda kv: kv[1]):
                ratio = float(name.strip("%")) / 100
                dist_pct = (price - current_price) / current_price * 100 if current_price else float("nan")
                fib_rows.append({
                    "Level": name,
                    "Ratio": round(ratio, 3),
                    "Price": round(price, 2),
                    "Distance from CMP %": round(dist_pct, 2),
                })
            fib_table = pd.DataFrame(fib_rows).style.set_properties(**{
                "background-color": COLOR_TABLE_BG,
                "color": COLOR_TABLE_TEXT,
                "border": f"1px solid {COLOR_TABLE_GRID}",
            })
            st.dataframe(fib_table, width="stretch", hide_index=True)
            st.caption(
                "**Ratio**: the Fibonacci retracement ratio for that level (0.0 = swing "
                "high, 1.0 = swing low). **Distance from CMP %**: positive = level is "
                "above the current price (potential resistance), negative = below "
                "(potential support)."
            )

            # --- Current position + likely next move ---------------------------------
            st.markdown("**Current Position & Likely Next Move**")
            st.markdown(
                f"Price (₹{current_price:.2f}) sits at the **{current_ratio * 100:.1f}%** "
                f"retracement of the swing range, nearest to the **{level_name}** level "
                f"(₹{level_price:.2f}), {distance_pct * 100:.2f}% away."
            )

            levels_above = [(n, p) for n, p in levels.items() if p > current_price]
            levels_below = [(n, p) for n, p in levels.items() if p < current_price]
            next_resistance = min(levels_above, key=lambda kv: kv[1]) if levels_above else None
            next_support = max(levels_below, key=lambda kv: kv[1]) if levels_below else None

            if next_resistance:
                r_name, r_price = next_resistance
                st.markdown(
                    f"- **Resistance above:** {r_name} level at ₹{r_price:.2f} "
                    f"({(r_price - current_price) / current_price * 100:.2f}% above CMP)."
                )
            if next_support:
                s_name, s_price = next_support
                st.markdown(
                    f"- **Support below:** {s_name} level at ₹{s_price:.2f} "
                    f"({(current_price - s_price) / current_price * 100:.2f}% below CMP)."
                )

            if hist_now >= 0:
                st.markdown(
                    "- MACD momentum is currently **bullish** (histogram above zero), "
                    "so price is more likely to move toward resistance / the swing "
                    "high while this holds."
                )
            else:
                st.markdown(
                    "- MACD momentum is currently **bearish** (histogram below zero), "
                    "so price is more likely to move toward support / the swing low "
                    "while this holds."
                )
    else:
        st.info("Run a scan to enable the Fibonacci retracement analysis.")

# ---------------------------------------------------------------------------
# Tab 4: Custom Analysis - on-demand AI Entry/Stop-Loss/Take-Profit write-up
# for any stocks picked from the scan universe, not just the daily shortlist.
# ---------------------------------------------------------------------------
if can_use_admin_tools:
    with tab_custom:
        st.subheader("🎯 Custom Stock Analysis")
        st.write(
            f"Get an AI-generated Entry / Stop-Loss / Take-Profit write-up for any "
            f"{config.CUSTOM_ANALYSIS_MIN_TICKERS}–{config.AI_TOP_PICKS_COUNT} stocks "
            "from the scan universe — useful for setups outside the daily shortlist. "
            f"Select at least {config.CUSTOM_ANALYSIS_MIN_TICKERS} stocks to run an analysis."
        )

        custom_selection = st.multiselect(
            "Select stocks to analyze",
            options=sorted(config.SCAN_UNIVERSE),
            max_selections=config.AI_TOP_PICKS_COUNT,
            key="custom_analysis_tickers",
        )

        n_selected = len(custom_selection)
        if n_selected < config.CUSTOM_ANALYSIS_MIN_TICKERS:
            st.info(
                f"Select at least {config.CUSTOM_ANALYSIS_MIN_TICKERS} stocks "
                f"({n_selected} selected so far)."
            )
        elif st.button("🤖 Get AI Analysis", type="primary", key="run_custom_analysis"):
            st.session_state["_pending_active_tab"] = st.session_state.get("active_tab")
            st.session_state["custom_analysis_in_progress"] = True
            st.session_state["custom_analysis_request"] = custom_selection
            st.rerun()

        if "custom_analysis_error" in st.session_state:
            st.error(st.session_state.pop("custom_analysis_error"))

        past_analyses = db_handler.get_available_custom_analyses()
        if past_analyses:
            st.markdown("---")

            _history_search = st.text_input(
                "🔍 Search past analyses by ticker",
                key="custom_history_search",
                placeholder="e.g. RELIANCE",
                label_visibility="collapsed",
            ).strip().upper()

            _filtered_analyses = [
                (aid, ts, tickers) for aid, ts, tickers in past_analyses
                if not _history_search or any(_history_search in t.upper() for t in tickers)
            ]

            if not _filtered_analyses:
                st.info("No past analyses match that ticker.")
            else:
                _analysis_labels = {}
                for _aid, _ts, _tickers in _filtered_analyses:
                    _names = ", ".join(t.replace(".NS", "") for t in _tickers[:4])
                    _suffix = f" +{len(_tickers) - 4} more" if len(_tickers) > 4 else ""
                    _analysis_labels[_aid] = f"{_ts} — {_names}{_suffix}"

                _sel_col, _del_col = st.columns([8, 1])
                with _sel_col:
                    selected_analysis_id = st.selectbox(
                        "Viewing analysis from:",
                        [a[0] for a in _filtered_analyses],
                        format_func=lambda a_id: _analysis_labels.get(a_id, str(a_id)),
                        key="custom_analysis_select",
                    )
                with _del_col:
                    st.markdown("<div style='padding-top:28px'>", unsafe_allow_html=True)
                    if st.button("🗑️ Delete", key="delete_custom_analysis", help="Delete this analysis"):
                        try:
                            db_handler.delete_custom_analysis(selected_analysis_id)
                            st.session_state.pop("custom_analysis_select", None)
                            st.rerun()
                        except Exception:
                            st.error("Could not delete — please try again.")
                    st.markdown("</div>", unsafe_allow_html=True)

                _shimmer_html = """
                    <div style="padding:4px 0">
                        <div class="shimmer-line" style="height:18px;width:45%;"></div>
                        <div class="shimmer-line" style="height:13px;width:100%;"></div>
                        <div class="shimmer-line" style="height:13px;width:96%;"></div>
                        <div class="shimmer-line" style="height:13px;width:89%;"></div>
                        <div class="shimmer-line" style="height:13px;width:93%;"></div>
                        <div class="shimmer-line" style="height:13px;width:72%;"></div>
                        <div class="shimmer-line" style="height:13px;width:98%;"></div>
                        <div class="shimmer-line" style="height:13px;width:81%;"></div>
                    </div>
                """
                _result_area = st.empty()
                _result_area.markdown(_shimmer_html, unsafe_allow_html=True)
                try:
                    result = db_handler.get_custom_analysis_by_id(selected_analysis_id)
                    if result:
                        result_tickers, result_df, result_commentary = result
                        with _result_area.container():
                            st.markdown(f"**Results for:** {', '.join(result_tickers)}")
                            st.markdown(result_commentary if result_commentary else "_No commentary available._")
                    else:
                        _result_area.empty()
                except Exception:
                    _result_area.empty()
                    st.error("Could not load analysis — please try again.")

# ---------------------------------------------------------------------------
# Tab 5: Admin - manage who can access this app (admins only)
# ---------------------------------------------------------------------------
if is_admin:
    with tab_admin:
        st.subheader("👑 User Access Management")

        users = db_handler.get_all_authorized_users()
        active_count = sum(1 for u in users if u["status"] == "active")
        st.metric("Registered users", f"{active_count} / {config.AUTH_MAX_USERS}")
        admin_list = ", ".join(sorted(config.AUTH_ADMIN_EMAILS)) or "none configured"
        st.write(
            "Users below were auto-registered on first Google sign-in "
            f"(first-come, first-served, capped at {config.AUTH_MAX_USERS}). "
            f"Admin account(s) ({admin_list}) always have access, don't count "
            "toward this limit, and are managed via the AUTH_ADMIN_EMAILS "
            "setting in .env, not from this table. **Revoke** signs a user out "
            "and frees their slot; **Restore** lets them back in."
        )

        if not users:
            st.info("No users have signed in yet.")
        else:
            h1, h2, h3, h4, h5 = st.columns([3, 3, 2, 1, 1])
            h1.markdown("**Email**")
            h2.markdown("**Name**")
            h3.markdown("**First login**")
            h4.markdown("**Status**")
            h5.markdown("**Action**")
            for u in users:
                c1, c2, c3, c4, c5 = st.columns([3, 3, 2, 1, 1])
                c1.write(u["email"])
                c2.write(u["name"] or "—")
                c3.write(u["first_login"])
                if u["status"] == "active":
                    c4.write("🟢 Active")
                    if c5.button("Revoke", key=f"revoke_{u['email']}"):
                        db_handler.revoke_user(u["email"])
                        st.rerun()
                else:
                    c4.write("🔴 Revoked")
                    if c5.button("Restore", key=f"restore_{u['email']}"):
                        db_handler.restore_user(u["email"])
                        st.rerun()

# ---------------------------------------------------------------------------
# Footer
# ---------------------------------------------------------------------------
st.markdown("---")
st.caption(config.FOOTER_ABOUT)
st.caption(config.FOOTER_DISCLAIMER)
